import os, shutil, sys
from pathlib import Path
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

source=Path(sys.argv[1]); output=Path(sys.argv[2]); output.parent.mkdir(parents=True,exist_ok=True)
if source.suffix.lower()=='.pdf': shutil.copyfile(source,output); raise SystemExit(0)
font='Helvetica'; font_path=Path('C:/Windows/Fonts/malgun.ttf')
if font_path.exists(): pdfmetrics.registerFont(TTFont('Malgun',str(font_path))); font='Malgun'
styles=getSampleStyleSheet(); styles['Title'].fontName=font; styles['BodyText'].fontName=font
story=[Paragraph(source.name,styles['Title']),Spacer(1,12)]
if source.suffix.lower() in ('.jpg','.jpeg','.png','.webp','.bmp'):
    from PIL import Image as PILImage
    with PILImage.open(source) as im: width,height=im.size
    max_w,max_h=A4[0]-50,A4[1]-90; scale=min(max_w/width,max_h/height)
    story.append(Image(str(source),width=width*scale,height=height*scale))
else:
    from openpyxl import load_workbook
    book=load_workbook(source,data_only=True,read_only=True)
    for sheet in book.worksheets:
        story.extend([Paragraph(sheet.title,styles['Heading2']),Spacer(1,5)])
        rows=[]
        for row in sheet.iter_rows(values_only=True):
            values=[str(value or '')[:80] for value in row]
            if any(values): rows.append(values)
            if len(rows)>=300: break
        if rows:
            columns=max(len(row) for row in rows); rows=[row+['']*(columns-len(row)) for row in rows]
            table=Table(rows,repeatRows=1,colWidths=[min(110,(A4[0]-50)/columns)]*columns)
            table.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),font),('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),.3,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#e8eef7')),('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(-1,-1),3),('RIGHTPADDING',(0,0),(-1,-1),3)]));story.extend([table,Spacer(1,12)])
SimpleDocTemplate(str(output),pagesize=A4,rightMargin=25,leftMargin=25,topMargin=25,bottomMargin=25,title=source.name).build(story)
